import openpyxl
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import os
import time

# Load workbook
workbook = openpyxl.load_workbook('attendance.xlsx')

# Select active worksheet
worksheet = workbook.active

# Global Variables
total_students = 0
absent_students = 0
attendance_threshold = 0.75
email_list = []


def save_excel():
    """
    This function saves the updated excel sheet.
    """
    workbook.save('attendance.xlsx')


def track_attendance():
    """
    This function tracks the attendance and sends an email alert if the attendance falls below the threshold.
    """
    global total_students, absent_students, email_list

    total_students = worksheet.max_row
    absent_students = worksheet.max_column

    attendance_ratio = (total_students - absent_students) / total_students

    if attendance_ratio < attendance_threshold:
        print(f"Attendance ratio is {attendance_ratio * 100}%. Sending an email alert.")
        for email in email_list:
            send_email(email)
    else:
        print(f"Attendance ratio is {attendance_ratio * 100}%. No action required.")


def send_email(to_email):
    """
    This function sends an email alert.
    """
    # Define email credentials and content
    smtp_server = "smtp.gmail.com"
    port = 587
    sender_email = "sender_email@gmail.com"
    password = "password"
    subject = "Attendance Alert"
    body = "Dear Admin, \n\nThe attendance ratio is below the threshold. \n\nBest Regards, \nThe Attendance Tracker"

    # Set up the SMTP server
    server = smtplib.SMTP(smtp_server, port)
    server.ehlo() # Can be omitted
    server.starttls() # Secure the connection
    server.login(sender_email, password)

    # Create the message
    msg = MIMEMultipart()

    # setup the parameters of the message
    msg['From'] = sender_email
    msg['To'] = to_email
    msg['Subject'] = subject

    # add in the message body
    msg.attach(MIMEText(body, 'plain'))

    # send the message via the server.
    server.sendmail(sender_email, to_email, msg.as_string())

    server.close()

    print(f"Email sent to {to_email}")


if __name__ == "__main__":
    while True:
       

        # Track attendance
        track_attendance()

        # Wait for a while before the next update
        time.sleep(3600) # 1 hour